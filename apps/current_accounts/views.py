from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
import csv

from .models import CurrentAccount, Transaction, Product, StockMovement, Warehouse


def to_decimal(value, default="0"):
    """Formdan gelen sayı metnini güvenli şekilde Decimal'e çevirir (virgül/nokta farkına tolerant)."""
    if value is None or value == "":
        value = default
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return Decimal(default)


# ---- STOK TAKİP ----

@login_required
def stock_tracking(request):
    products = Product.objects.filter(user=request.user)
    total_stock_value = sum((p.stock_value for p in products), Decimal("0"))
    low_stock_count = sum(1 for p in products if p.is_low_stock)
    return render(request, 'current_accounts/stock-tracking.html', {
        'products': products,
        'total_stock_value': total_stock_value,
        'low_stock_count': low_stock_count,
    })


@login_required
def product_save(request, id=None):
    product = get_object_or_404(Product, id=id, user=request.user) if id else Product(user=request.user)
    if request.method == "POST":
        product.code = request.POST.get("code", "")
        product.name = request.POST.get("name")
        product.category = request.POST.get("category", "")
        product.unit = request.POST.get("unit", "Adet")
        product.quantity = to_decimal(request.POST.get("quantity"), "0")
        product.unit_price = to_decimal(request.POST.get("unit_price"), "0")
        product.cost_price = to_decimal(request.POST.get("cost_price"), "0")
        product.vat_rate = to_decimal(request.POST.get("vat_rate"), "20")
        product.min_stock_level = to_decimal(request.POST.get("min_stock_level"), "0")
        product.save()
    return redirect("stock_tracking")


@login_required
def product_delete(request, id):
    get_object_or_404(Product, id=id, user=request.user).delete()
    return redirect("stock_tracking")


# ---- STOK HAREKETİ (Aşama 20) ----

@login_required
def stock_movement_list(request):
    movements = StockMovement.objects.filter(user=request.user).select_related("product")
    products = Product.objects.filter(user=request.user, is_active=True)
    return render(request, 'current_accounts/stock-movement.html', {
        'movements': movements,
        'products': products,
    })


@login_required
def stock_movement_save(request):
    if request.method != "POST":
        return redirect("stock_movement_list")

    product = get_object_or_404(Product, id=request.POST.get("product"), user=request.user)
    movement_type = request.POST.get("type")
    quantity = to_decimal(request.POST.get("quantity"), "0")
    date_str = request.POST.get("date") or datetime.now().date().isoformat()

    if movement_type not in dict(StockMovement.TYPE_CHOICES):
        messages.error(request, "Geçersiz hareket türü.")
        return redirect("stock_movement_list")

    if quantity <= 0:
        messages.error(request, "Miktar sıfırdan büyük olmalıdır.")
        return redirect("stock_movement_list")

    # ÖNEMLİ: Çıkış hareketi, kaydedilmeden ÖNCE mevcut stoğu aşıp aşmadığı
    # kontrol edilir (negatif stoğa düşmeyi engeller). Sinyal yalnızca
    # buradan geçmiş, doğrulanmış hareketleri Product.quantity'e yansıtır.
    if movement_type == "out" and quantity > product.quantity:
        messages.error(
            request,
            f"Yetersiz stok: '{product.name}' için mevcut {product.quantity} {product.unit}, "
            f"talep edilen {quantity} {product.unit}."
        )
        return redirect("stock_movement_list")

    StockMovement.objects.create(
        user=request.user, product=product, type=movement_type,
        reason=request.POST.get("reason", "other"),
        quantity=quantity, date=date_str,
        reference_note=request.POST.get("reference_note", ""),
    )
    messages.success(request, f"{product.name} için {dict(StockMovement.TYPE_CHOICES)[movement_type].lower()} hareketi kaydedildi.")
    return redirect("stock_movement_list")


# ---- DEPO (Aşama 19) ----

@login_required
def warehouse_list(request):
    warehouses = Warehouse.objects.filter(user=request.user)
    return render(request, 'current_accounts/warehouse-list.html', {"warehouses": warehouses})


@login_required
def warehouse_save(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            Warehouse.objects.create(user=request.user, name=name, location=request.POST.get("location", ""))
        else:
            messages.error(request, "Depo adı zorunludur.")
    return redirect("warehouse_list")


@login_required
def warehouse_delete(request, id):
    get_object_or_404(Warehouse, id=id, user=request.user).delete()
    return redirect("warehouse_list")


# ---- CARİ BAKİYE ----

@login_required
def current_balance(request):
    accounts = CurrentAccount.objects.filter(user=request.user)
    total_receivable = sum((a.balance for a in accounts if a.balance > 0), Decimal("0"))
    total_payable = sum((-a.balance for a in accounts if a.balance < 0), Decimal("0"))
    return render(request, 'current_accounts/current-balance.html', {
        'accounts': accounts,
        'total_receivable': total_receivable,
        'total_payable': total_payable,
    })


@login_required
def account_save(request, id=None):
    account = get_object_or_404(CurrentAccount, id=id, user=request.user) if id else CurrentAccount(user=request.user)
    if request.method == "POST":
        account.type = request.POST.get("type", "customer")
        account.name = request.POST.get("name")
        account.company_name = request.POST.get("company_name", "")
        account.tax_id = request.POST.get("tax_id", "")
        account.tax_office = request.POST.get("tax_office", "")
        account.email = request.POST.get("email", "")
        account.phone = request.POST.get("phone", "")
        account.address = request.POST.get("address", "")
        account.credit_limit = to_decimal(request.POST.get("credit_limit"), "0")
        account.notes = request.POST.get("notes", "")
        if not id:
            account.balance = to_decimal(request.POST.get("balance"), "0")
        account.save()
    return redirect("current_balance")


@login_required
def account_delete(request, id):
    get_object_or_404(CurrentAccount, id=id, user=request.user).delete()
    return redirect("current_balance")


# ---- CARİ EKSTRE ----

def _build_statement_rows(account):
    """Bir carinin işlemlerini, kronolojik sırayla ve koşan (running) bakiyeyle döner."""
    transactions = account.transactions.order_by("date", "id")
    balance = Decimal("0")
    rows = []
    for t in transactions:
        if t.type == "debit":
            balance += t.amount
        else:
            balance -= t.amount
        rows.append({"t": t, "running_balance": balance})
    return rows


def _aging_buckets(account):
    """
    Aşama 17 notu: "Cari Hesap Yaşlandırma ve Bakiye Raporları."

    Basitleştirilmiş yaklaşım: tam FIFO eşleştirmesi (hangi borcun hangi
    tahsilatla kapandığı) yerine, henüz ödenmemiş TOPLAM borç tutarını
    borç kayıtlarının yaşına göre 4 kovaya (0-30 / 31-60 / 61-90 / 90+ gün)
    dağıtıyoruz. Küçük ve anlaşılır bir yaklaşım; tam FIFO eşleştirmesi
    ayrı, daha büyük bir görev olarak ele alınmalı.
    """
    today = date.today()
    buckets = {"d0_30": Decimal("0"), "d31_60": Decimal("0"), "d61_90": Decimal("0"), "d90_plus": Decimal("0")}
    debit_total = Decimal("0")
    credit_total = Decimal("0")

    for t in account.transactions.all():
        if t.type == "debit":
            debit_total += t.amount
        else:
            credit_total += t.amount

    # Toplam tahsil edilmemiş bakiye (basit netleştirme):
    outstanding = debit_total - credit_total
    if outstanding <= 0:
        return buckets  # borç yok / fazla ödeme var

    # Outstanding tutarı, en eski borç kayıtlarından başlayarak yaş kovalarına dağıt.
    remaining = outstanding
    for t in account.transactions.filter(type="debit").order_by("date"):
        if remaining <= 0:
            break
        take = min(t.amount, remaining)
        age_days = (today - t.date).days
        if age_days <= 30:
            buckets["d0_30"] += take
        elif age_days <= 60:
            buckets["d31_60"] += take
        elif age_days <= 90:
            buckets["d61_90"] += take
        else:
            buckets["d90_plus"] += take
        remaining -= take

    return buckets


@login_required
def account_statement(request, id=None):
    accounts = CurrentAccount.objects.filter(user=request.user)
    selected_account = None
    running_rows = []
    aging = None

    account_id = id or request.GET.get("account_id")
    if account_id:
        selected_account = get_object_or_404(CurrentAccount, id=account_id, user=request.user)
        running_rows = _build_statement_rows(selected_account)
        aging = _aging_buckets(selected_account)

    return render(request, 'current_accounts/account-statement.html', {
        'accounts': accounts,
        'selected_account': selected_account,
        'running_rows': running_rows,
        'aging': aging,
    })


@login_required
def account_statement_export(request, id, file_format):
    """
    Cari ekstreyi CSV ya da Excel (.xlsx) olarak indirir (Aşama 16).

    Not: PDF çıktısı bu pakete dahil edilmedi — reportlab/weasyprint gibi ek bir
    bağımlılık gerektiriyor ve mevcut requirements.txt'de yok; CSV/XLSX ile aynı
    prensipte (aynı _build_statement_rows verisi) kolayca eklenebilir.
    """
    account = get_object_or_404(CurrentAccount, id=id, user=request.user)
    rows = _build_statement_rows(account)
    safe_name = "".join(ch for ch in account.name if ch.isalnum() or ch in (" ", "-", "_")).strip() or "cari"

    if file_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = f'attachment; filename="ekstre_{safe_name}.csv"'
        writer = csv.writer(response)
        writer.writerow(["Tarih", "Açıklama", "Belge No", "Borç", "Alacak", "Bakiye"])
        for row in rows:
            t = row["t"]
            writer.writerow([
                t.date.strftime("%d.%m.%Y"),
                t.description,
                t.document_number,
                t.amount if t.type == "debit" else "",
                t.amount if t.type == "credit" else "",
                row["running_balance"],
            ])
        return response

    if file_format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Cari Ekstre"

        ws.merge_cells("A1:F1")
        ws["A1"] = f"{account.name} - Cari Hesap Ekstresi"
        ws["A1"].font = Font(bold=True, size=13)

        headers = ["Tarih", "Açıklama", "Belge No", "Borç", "Alacak", "Bakiye"]
        header_fill = PatternFill("solid", fgColor="1F3864")
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        r = 3
        for row in rows:
            r += 1
            t = row["t"]
            ws.cell(row=r, column=1, value=t.date.strftime("%d.%m.%Y"))
            ws.cell(row=r, column=2, value=t.description)
            ws.cell(row=r, column=3, value=t.document_number)
            ws.cell(row=r, column=4, value=float(t.amount) if t.type == "debit" else None)
            ws.cell(row=r, column=5, value=float(t.amount) if t.type == "credit" else None)
            ws.cell(row=r, column=6, value=float(row["running_balance"]))

        widths = [14, 40, 16, 14, 14, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="ekstre_{safe_name}.xlsx"'
        wb.save(response)
        return response

    return HttpResponse("Desteklenmeyen format. 'csv' ya da 'xlsx' kullanın.", status=400)


@login_required
def transaction_save(request, account_id):
    account = get_object_or_404(CurrentAccount, id=account_id, user=request.user)
    if request.method == "POST":
        transaction = Transaction.objects.create(
            current_account=account,
            type=request.POST.get("type", "debit"),
            amount=to_decimal(request.POST.get("amount"), "0"),
            description=request.POST.get("description", ""),
            date=request.POST.get("date"),
            document_number=request.POST.get("document_number", ""),
        )
        if transaction.type == "debit":
            account.balance += transaction.amount
        else:
            account.balance -= transaction.amount
        account.save(update_fields=["balance"])
    return redirect(f"/current-accounts/cari-ekstre/{account_id}/")


@login_required
def transaction_delete(request, id):
    transaction = get_object_or_404(Transaction, id=id, current_account__user=request.user)
    account = transaction.current_account
    if transaction.type == "debit":
        account.balance -= transaction.amount
    else:
        account.balance += transaction.amount
    account.save(update_fields=["balance"])
    account_id = account.id
    transaction.delete()
    return redirect(f"/current-accounts/cari-ekstre/{account_id}/")


# ---- ALINAN / GİDEN E-FATURALAR ----
# Bu veriler zaten e-Fatura modülünde (Gelen/Giden E-Faturalar) tutuluyor.
# Tekrar aynı tabloyu burada oluşturmak yerine oraya yönlendiriyoruz;
# böylece tek bir doğru veri kaynağı (single source of truth) kalıyor.

@login_required
def received_invoices(request):
    return redirect("invoices_incoming")


@login_required
def sent_invoices(request):
    return redirect("invoices_sent")
